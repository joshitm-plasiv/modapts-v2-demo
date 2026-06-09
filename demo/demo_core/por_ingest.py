"""
POR ingestion — parse the client POR workbook into a structured PlanOfRecord.

The workbook is a consistent 5-sheet template:
  POR_ProgramInputs       Parameter | Value | Units            (program demand)
  POR_ProcessInputs       Process   | Line                     (the lines in the plant)
  POR_LineTargets         per-line targets, sequence, manning, rate
  POR_ProcessFlow         per-station: cycle time, yield, downtime, qty, cost, …
  POR_ProcessActivityData per-activity within a station (the manual ones are MODAPTS work)

Parsing is DETERMINISTIC and matches columns by header NAME (substring), so minor
column shifts between clients don't break it. The Excel has no per-field provenance,
so every value is tagged source="POR document" — we surface that, we don't invent
confidence. PDF ingestion (text+tables) maps into the same structure; added separately.
"""
from __future__ import annotations

from dataclasses import dataclass, field, asdict
from typing import Any, Optional

from openpyxl import load_workbook

PROVENANCE = "POR document"   # the workbook states facts; no measured/estimate split


# ── helpers ──────────────────────────────────────────────────────────────────────
def _norm(s: Any) -> str:
    return str(s).strip().lower() if s is not None else ""


def _num(v: Any) -> Optional[float]:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except (TypeError, ValueError):
        return None


def _truthy(v: Any) -> bool:
    return _norm(v) in ("y", "yes", "true", "1", "x")


def _read_table(ws) -> tuple[list[str], list[dict]]:
    """Header = first row; data rows follow until the first column goes empty."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [("" if h is None else str(h).strip()) for h in rows[0]]
    out: list[dict] = []
    for r in rows[1:]:
        if not r or r[0] is None or (isinstance(r[0], str) and not r[0].strip()):
            break  # blank key column → end of data (sheets are padded to 1000 rows)
        rec = {headers[i]: r[i] for i in range(min(len(headers), len(r))) if headers[i]}
        out.append(rec)
    return headers, out


def _pick(rec: dict, *aliases: str):
    """Return the first value whose header contains any alias (case-insensitive)."""
    for key, val in rec.items():
        k = _norm(key)
        if any(a in k for a in aliases):
            return val
    return None


# ── structured POR ─────────────────────────────────────────────────────────────
@dataclass
class Activity:
    line: str
    station: str
    activity_type: str          # MANUAL | AUTOMATIC | …
    precedence: Optional[int]
    name: str
    lean_category: str
    lean_classification: str    # VA | RNVA | NVA
    frequency: Optional[float]
    activity_time_s: Optional[float]
    parallel: bool
    raw: dict = field(default_factory=dict)

    @property
    def is_manual(self) -> bool:
        return _norm(self.activity_type) == "manual"


@dataclass
class Station:
    line: str
    station_id: str
    name: str
    station_type: str
    cycle_time_s: Optional[float]
    parts_per_cycle: Optional[float]
    yield_rate: Optional[float]
    retest_rate: Optional[float]
    planned_downtime: Optional[float]
    unplanned_downtime: Optional[float]
    allowance: Optional[float]
    offline: bool
    machine_qty: Optional[float]
    operator_qty: Optional[float]
    cost_per_fixture: Optional[float]
    cost_per_machine: Optional[float]
    activities: list[Activity] = field(default_factory=list)
    raw: dict = field(default_factory=dict)

    @property
    def manual_activity_time_s(self) -> float:
        return sum(a.activity_time_s or 0 for a in self.activities if a.is_manual)


@dataclass
class Line:
    name: str
    process: str = ""
    target_throughput: Optional[float] = None
    throughput_unit: str = ""
    working_hours_per_day: Optional[float] = None
    operator_hourly_rate: Optional[float] = None
    preceding_line: str = ""
    succeeding_line: str = ""
    planned_num_lines: Optional[float] = None
    stations: list[Station] = field(default_factory=list)
    raw: dict = field(default_factory=dict)

    @property
    def bottleneck(self) -> Optional[Station]:
        timed = [s for s in self.stations if s.cycle_time_s]
        return max(timed, key=lambda s: s.cycle_time_s) if timed else None


@dataclass
class PlanOfRecord:
    source_file: str
    program: dict = field(default_factory=dict)     # {param: {"value":…, "units":…}}
    lines: list[Line] = field(default_factory=list)
    provenance: str = PROVENANCE

    def line_names(self) -> list[str]:
        return [ln.name for ln in self.lines]

    def get_line(self, name: str) -> Optional[Line]:
        n = _norm(name)
        for ln in self.lines:
            if _norm(ln.name) == n or n in _norm(ln.name):
                return ln
        return None

    def all_stations(self) -> list[Station]:
        return [s for ln in self.lines for s in ln.stations]

    def summary(self) -> dict:
        return {
            "source_file": self.source_file,
            "provenance": self.provenance,
            "program": self.program,
            "lines": len(self.lines),
            "stations": len(self.all_stations()),
            "activities": sum(len(s.activities) for s in self.all_stations()),
            "manual_activities": sum(
                1 for s in self.all_stations() for a in s.activities if a.is_manual),
            "line_detail": [
                {"line": ln.name, "process": ln.process,
                 "stations": len(ln.stations),
                 "target_throughput": ln.target_throughput,
                 "throughput_unit": ln.throughput_unit,
                 "bottleneck": (ln.bottleneck.station_id if ln.bottleneck else None),
                 "bottleneck_ct_s": (ln.bottleneck.cycle_time_s if ln.bottleneck else None)}
                for ln in self.lines
            ],
        }


# ── parser ───────────────────────────────────────────────────────────────────────
def _assemble(prog_rows, proc_rows, lt_rows, sf_rows, ad_rows, source_file) -> PlanOfRecord:
    """Build a PlanOfRecord from the five row-lists (shared by xlsx and pdf)."""
    program: dict = {}
    for rec in prog_rows:
        param = _pick(rec, "parameter")
        if param is None:
            continue
        v = _pick(rec, "value")
        program[str(param).strip()] = {"value": _num(v) if _num(v) is not None else v,
                                        "units": _pick(rec, "unit")}

    process_of: dict[str, str] = {}
    line_order: list[str] = []
    for rec in proc_rows:
        line = _pick(rec, "line")
        if line:
            line = str(line).strip()
            process_of[_norm(line)] = str(_pick(rec, "process") or "").strip()
            if line not in line_order:
                line_order.append(line)

    line_targets: dict[str, dict] = {}
    for rec in lt_rows:
        line = _pick(rec, "line")
        if not line:
            continue
        line = str(line).strip()
        if line not in line_order:
            line_order.append(line)
        line_targets[_norm(line)] = rec

    stations_by_line: dict[str, list[Station]] = {}
    for rec in sf_rows:
        line = str(_pick(rec, "line") or "").strip()
        sid = str(_pick(rec, "station id", "station_id") or _pick(rec, "station") or "").strip()
        if not line or not sid:
            continue
        stations_by_line.setdefault(_norm(line), []).append(Station(
            line=line, station_id=sid,
            name=str(_pick(rec, "station name", "name") or "").strip(),
            station_type=str(_pick(rec, "station type", "type") or "").strip(),
            cycle_time_s=_num(_pick(rec, "cycle time")),
            parts_per_cycle=_num(_pick(rec, "parts per cycle", "parts/cycle")),
            yield_rate=_num(_pick(rec, "yield")),
            retest_rate=_num(_pick(rec, "retest")),
            planned_downtime=_num(_pick(rec, "planned downtime")),
            unplanned_downtime=_num(_pick(rec, "unplanned downtime")),
            allowance=_num(_pick(rec, "allowance")),
            offline=_truthy(_pick(rec, "offline")),
            machine_qty=_num(_pick(rec, "machine qty", "planned machine")),
            operator_qty=_num(_pick(rec, "operator qty", "planned operator")),
            cost_per_fixture=_num(_pick(rec, "cost per fixture")),
            cost_per_machine=_num(_pick(rec, "cost per machine")),
            raw=rec,
        ))

    for rec in ad_rows:
        line = str(_pick(rec, "line") or "").strip()
        sid = str(_pick(rec, "station") or "").strip()
        if not line or not sid:
            continue
        prec = _num(_pick(rec, "precedence"))
        act = Activity(
            line=line, station=sid,
            activity_type=str(_pick(rec, "activity type") or "").strip(),
            precedence=(int(prec) if prec is not None else None),
            name=str(_pick(rec, "activity name", "name") or "").strip(),
            lean_category=str(_pick(rec, "lean category") or "").strip(),
            lean_classification=str(_pick(rec, "lean classification") or "").strip(),
            frequency=_num(_pick(rec, "frequency")),
            activity_time_s=_num(_pick(rec, "activity time")),
            parallel=_truthy(_pick(rec, "parallel")),
            raw=rec,
        )
        for st in stations_by_line.get(_norm(line), []):
            if _norm(st.station_id) == _norm(sid):
                st.activities.append(act)
                break

    lines: list[Line] = []
    for name in line_order:
        lt = line_targets.get(_norm(name), {})
        lines.append(Line(
            name=name, process=process_of.get(_norm(name), ""),
            target_throughput=_num(_pick(lt, "target throughput") if lt else None),
            throughput_unit=str(_pick(lt, "throughput unit", "unit") or "").strip() if lt else "",
            working_hours_per_day=_num(_pick(lt, "working hours", "hours per") if lt else None),
            operator_hourly_rate=_num(_pick(lt, "hourly rate", "operator hourly") if lt else None),
            preceding_line=str(_pick(lt, "preceding line") or "").strip() if lt else "",
            succeeding_line=str(_pick(lt, "succeeding line") or "").strip() if lt else "",
            planned_num_lines=_num(_pick(lt, "planned num lines", "num lines") if lt else None),
            stations=sorted(stations_by_line.get(_norm(name), []), key=lambda s: s.station_id),
            raw=lt,
        ))
    return PlanOfRecord(source_file=source_file, program=program, lines=lines)


def load_por_xlsx(path: str) -> PlanOfRecord:
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = {ws.title: ws for ws in wb.worksheets}

    def rows_of(name: str) -> list[dict]:
        for title, ws in sheets.items():
            if _norm(name) in _norm(title):
                return _read_table(ws)[1]
        return []

    por = _assemble(rows_of("ProgramInputs"), rows_of("ProcessInputs"),
                    rows_of("LineTargets"), rows_of("ProcessFlow"),
                    rows_of("ProcessActivityData"), path.split("/")[-1])
    wb.close()
    return por


# Header signatures used to route a PDF table to the right sheet bucket.
def _classify_table(headers: list[str]) -> Optional[str]:
    h = " | ".join(_norm(x) for x in headers)
    if "parameter" in h and "value" in h:
        return "program"
    if "activity" in h and ("precedence" in h or "activity time" in h):
        return "activity"
    if "station" in h and ("cycle time" in h or "station name" in h or "station id" in h):
        return "flow"
    if "line" in h and "target throughput" in h:
        return "targets"
    if "process" in h and "line" in h:
        return "process"
    return None


def load_por_pdf(path: str) -> PlanOfRecord:
    """PDF that is text + tables (not scanned). Extract every table, route each to its
    sheet bucket by header signature, then assemble with the same deterministic mapping.
    Tables may span pages — rows accumulate. (Verify on a real PDF before relying on it.)"""
    try:
        import pdfplumber
    except ImportError:
        raise NotImplementedError("pdfplumber not installed. Run: pip install pdfplumber")

    buckets: dict[str, list[dict]] = {k: [] for k in ("program", "process", "targets", "flow", "activity")}
    last_bucket: Optional[str] = None
    last_headers: list[str] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for tbl in (page.extract_tables() or []):
                if not tbl or len(tbl) < 2:
                    continue
                headers = [("" if c is None else str(c).strip()) for c in tbl[0]]
                bucket = _classify_table(headers)
                if bucket is None and last_bucket and len(headers) == len(last_headers):
                    bucket, headers, body = last_bucket, last_headers, tbl   # continuation
                else:
                    body = tbl[1:]
                if bucket is None:
                    continue
                for r in body:
                    rec = {headers[i]: r[i] for i in range(min(len(headers), len(r))) if headers[i]}
                    if any(v not in (None, "") for v in rec.values()):
                        buckets[bucket].append(rec)
                last_bucket, last_headers = bucket, headers

    return _assemble(buckets["program"], buckets["process"], buckets["targets"],
                     buckets["flow"], buckets["activity"], path.split("/")[-1])


def load_por(path: str) -> PlanOfRecord:
    """Dispatch by extension. xlsx is deterministic; pdf reuses the same mapping over
    extracted tables."""
    low = path.lower()
    if low.endswith((".xlsx", ".xlsm")):
        return load_por_xlsx(path)
    if low.endswith(".pdf"):
        return load_por_pdf(path)
    raise NotImplementedError(f"POR ingestion for .{path.split('.')[-1]} not supported "
                              f"(xlsx and pdf are).")
